import csv
import urllib3
import pandas as pd
from openpyxl import load_workbook
import time
import math

downloadAcf = True
downloadJobs = True
allIn = False
userKey = "H5KWP6N8GC"
sleepTime = 10
speedFactor = 0.8

checkAcf = []
checkAcf.append("BAe Jetstream 32")
checkAcf.append("Douglas DC-3")
checkAcf.append("ATR 42-500")
checkAcf.append("Saab 340B")
checkAcf.append("Dornier 328")
checkAcf.append("Embraer 120")
checkAcf.append("Dornier 228")
checkAcf.append("Bombardier CRJ-200ER")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")
#checkAcf.append("")

#Saab 340B
#ATR 72-500
#BAe 146-100 (Avro RJ70)
#Bombardier CRJ700-ER
#Bombardier Dash-8 Q400
#Bombardier Dash-8 Q300
#DeHavilland Dash 7
#Douglas DC-4
#Douglas DC-6
#Douglas DC-6B
#Douglas DC-6B (PMDG)
#Douglas DC-7B
#Douglas DC-7C
#Fairchild C119
#Fokker 50
#Lockheed C-130 (Capt Sim)
#Lockheed C-130 (Generic)
#Lockheed L049 (A2A)
#Lockheed P-3C (L-188)

if allIn == True:
    checkAcf.append("Airbus A320")

######################################################################################
def calc_distance_nm(lat1, long1, lat2, long2):
    lat1 = lat1 * math.pi / 180
    long1 = long1 * math.pi / 180
    lat2 = lat2 * math.pi / 180
    long2 = long2 * math.pi / 180

    rEarth = 6372.797

    dlat = lat2 - lat1
    dlong = long2 - long1

    x1 = math.sin(dlat / 2)
    x2 = math.cos(lat1)
    x3 = math.cos(lat2)
    x4 = math.sin(dlong / 2)

    x5 = x1 * x1
    x6 = x2 * x3 * x4 * x4

    temp1 = x5 + x6

    y1 = math.sqrt(temp1)
    y2 = math.sqrt(1.0 - temp1)

    temp2 = 2 * math.atan2(y1, y2)

    rangeKm = temp2 * rEarth

    CalcRange = rangeKm * 0.539957

    return CalcRange
######################################################################################
icao_list_icao = []
icao_list_lat = []
icao_list_lon = []
icao_list_name = []
######################################################################################
def get_icao_name(icao):
    icao_name = "xxx"
    for icao_index in range(0, len(icao_list_icao)):
        if (icao_list_icao[icao_index] == icao):
            icao_name = icao_list_name[icao_index]
    return icao_name
######################################################################################
def calc_distance_by_icao(icao1, icao2):
    distance = 0
    lat1 = -1000
    lat2 = -1000
    
    lon1 = -1000
    lon2 = -1000

    for icao_index in range(0, len(icao_list_icao)):
        if (icao_list_icao[icao_index] == icao1):
            lat1 = icao_list_lat[icao_index]
            lon1 = icao_list_lon[icao_index]
        if (icao_list_icao[icao_index] == icao2):    
            lat2 = icao_list_lat[icao_index]
            lon2 = icao_list_lon[icao_index]

    if (lat1 == -1000) or (lat2 == -1000) or (lon1 == -1000) or(lon2 == -1000):
        print("############################ICAO ERROR##############################")

    return calc_distance_nm(lat1, lon1, lat2, lon2)
######################################################################################



with open('icaodata.csv') as csv_file:
    reader = csv.reader(csv_file, delimiter=',')
    index = 0
    for row in reader:
        if index > 0:
            icao_list_icao.append(row[0])
            icao_list_lat.append(float(row[1]))
            icao_list_lon.append(float(row[2]))
            icao_list_name.append(row[5] + ", " + row[6] + ", " + row[7] + ", " + row[8])
        index +=1
    csv_file.close()


acfName = []
acfCrew = []
acfSeats = []
acfSpeed = []
acfFuel = []
acfGPH = []
acfPayload = []


# ACF: 0 Name, 1 Crew, 2 Seats, 3 Cruise, 4 Fuel, 5 Fuel Type, 6 GPH, 7 Payload, 8 Price

with open("acf.txt") as csv_file:
    reader = csv.reader(csv_file, delimiter=',')
    for row in reader:
        if (len(row) >= 7):
            acfName.append(row[0])
            acfCrew.append(float(row[1]))
            acfSeats.append(float(row[2]))
            acfSpeed.append(float(row[3]))
            acfFuel.append(float(row[4]))
            acfGPH.append(float(row[6]))
            acfPayload.append(float(row[7]))
			
pandasBook = load_workbook("output.xlsx")
pandasWriter = pd.ExcelWriter("output.xlsx", engine='openpyxl')
pandasWriter.book = pandasBook

for curName in checkAcf:
    acfIndex = -1
    for index in range(0, len(acfName)):
        if acfName[index] == curName:
            acfIndex = index

    if (acfIndex >= 0):
        print("Acf Index: " + str(acfIndex))
        acfFileName = acfName[acfIndex].replace(" ", "_") + ".csv"
        acfSheetName = acfName[acfIndex].replace(" ", "_")
        if (downloadAcf == True):
            time.sleep(sleepTime)
            acfUrlName = acfName[acfIndex].replace(" ", "%20")
            
            url = "http://server.fseconomy.net/data?userkey=" + userKey + "&format=csv&query=aircraft&search=makemodel&makemodel=" + acfUrlName
            
            http = urllib3.PoolManager()
            r = http.request('GET', url, preload_content = False)
            # acfDataUrl = urllib3.urlopen(url)
            acfBinData = r.read()
            with open(acfFileName, "wb") as acfFileHandle:
                acfFileHandle.write(acfBinData)
                print("File written: " + acfFileName)
            r.release_conn()
        
        listSerial = []
        listReg = []
        listLocation = []
        listLocationName = []
        listHome = []
        listRentalDry = []
        listRentalWet = []
        listBonus = []
        listCargoSpace = []

        rentDry = 0
        rentWet = 0
        rentedby = ""
        needsRepair = 0
        feeOwed = 0



        with open(acfFileName) as csv_file:
            index = 0
            reader = csv.reader(csv_file, delimiter=',')
            for row in reader:
                if (len(row) >= 23) and (index > 0):
                    rentDry = float(row[10])
                    rentWet = float(row[11])
                    rentedby = row[15]
                    needsRepair = float(row[17])
                    feeOwed = float(row[23])

                    if ((rentDry > 0) or (allIn == True)) and (rentedby == "Not rented.") and (needsRepair == 0) and (feeOwed == 0):
                        listSerial.append(int(row[0]))
                        listReg.append(row[2])
                        listLocation.append(row[4])
                        listLocationName.append(row[5])
                        listHome.append(row[6])
                        listRentalDry.append(float(row[10]))
                        listRentalWet.append(float(row[11]))
                        listBonus.append(float(row[13]))
                        fuel = acfFuel[acfIndex] * 2.73 * float(row[16])
                                
                        listCargoSpace.append(acfPayload[acfIndex] - fuel)

                index = index + 1

        print("Total Aircraft: " + str(len(listReg)))
        #pandasData = pd.DataFrame({  'Reg': listReg,
        #                'Location' : listLocation,
        #                'LocationName' : listLocationName,
        #                'Home' : listHome,
        #                'Dry' : listRentalDry,
        #                'Wet' : listRentalWet,
        #                'Bonus' : listBonus,
        #                'Cargo' : listCargoSpace
        #                })


        ##pandasWriter = pd.ExcelWriter("output.xlsx", engine='xlsxwriter')
        ##pandasBook = load_workbook("output.xlsx")
        #pandasWriter = pd.ExcelWriter("output.xlsx", engine='openpyxl')
        ##pandasWriter.book = pandasBook
        #pandasData.to_excel(pandasWriter, sheet_name = "mySheet4")
        #pandasWriter.save()

        icaoList = ""

        for index in range(0, len(listLocation)):
            if (icaoList != ""):
                icaoList += "-"
            icaoList += listLocation[index]

        print(icaoList)
        jobsFileName = "jobs_" + acfFileName
        if (downloadJobs == True):
            time.sleep(sleepTime)
            url = "http://server.fseconomy.net/data?userkey=" + userKey + "&format=csv&query=icao&search=jobsfrom&icaos=" + icaoList
            
            http = urllib3.PoolManager()
            r = http.request('GET', url, preload_content = False)
            # acfDataUrl = urllib3.urlopen(url)
            acfBinData = r.read()
            with open(jobsFileName, "wb") as acfFileHandle:
                acfFileHandle.write(acfBinData)
                print("File written: " + jobsFileName)
            r.release_conn()

        jobsFrom = []
        jobsTo = []
        jobsAmount = []
        jobsUnit = []
        jobsPay = []
        jobsGreen = []
        jobsType = []
        jobsAircraft = []

        jobsWeight = []
        jobsPayPerWeight = []

        with open(jobsFileName) as csv_file:
            index = 0
            reader = csv.reader(csv_file, delimiter=',')
            for row in reader:
                if (len(row) >= 13) and (index > 0):
                    jobsFrom.append(row[1])
                    jobsTo.append(row[2])
                    jobsAmount.append(float(row[4]))
                    jobsUnit.append(row[5])
                    jobsPay.append(float(row[7]))
                    if (row[11] == "true"): jobsGreen.append(True)
                    else:                   jobsGreen.append(False)
                    jobsType.append(row[12])
                    jobsAircraft.append(int(row[13]))

                    if (int(row[13]) != 0):
                        jobsWeight.append(0)
                        jobsPayPerWeight.append(0)
                    else:
                        weight = 77 * float(row[4])
                        if row[5] == "kg": weight = float(row[4])
                        if (weight == 0):    weight = 0.01
                        
                        jobsWeight.append(weight)
                        jobsPayPerWeight.append(float(row[7]) / weight )
                    
                index += 1

        selectedFrom = []
        selectedTo = []
        selectedAmount = []
        selectedUnit = []
        selectedPay = []
        selectedGreen = []
        selectedType = []
        selectedAircraft = []

        selectedWeight = []
        selectedDistance = []
        selectedBonusDistance = []
        selectedPayPerWeight = []
        selectedIncome = -1

        bestFrom = []
        bestTo = []
        bestAmount = []
        bestUnit = []
        bestPay = []
        bestGreen = []
        bestType = []
        bestAircraft = []

        bestWeight = []
        bestDistance = []
        bestBonusDistance = []
        bestPayPerWeight = []
        bestIncome = -1

        finalFrom = []
        finalTo = []
        finalFromName = []
        finalToName = []
        finalAmount = []
        finalUnit = []
        finalPay = []
        finalGreen = []
        finalType = []
        finalAircraft = []

        finalWeight = []
        finalDistance = []
        finalBonusDistance = []
        finalPayPerWeight = []
        finalCosts = []
        finalPayment = []
        finalIncome = []

        if (allIn == True):
            for curAcfIndex in range(0, len(listLocation)):
                print(acfSheetName + ": " + listLocation[curAcfIndex])
                for jobIndex in range(0, len(jobsFrom)):
                    if (jobsFrom[jobIndex] == listLocation[curAcfIndex]) and (jobsType[jobIndex] == "All-In") :
                        finalFrom.append(jobsFrom[jobIndex])
                        finalTo.append(jobsTo[jobIndex])
                        finalFromName.append(get_icao_name(jobsFrom[jobIndex]))
                        finalToName.append(get_icao_name(jobsTo[jobIndex]))
                        finalAmount.append("")
                        finalUnit.append("")
                        finalPay.append(jobsPay[jobIndex])
                        finalGreen.append("")
                        finalType.append("")
                        finalAircraft.append("")

                        finalWeight.append("")
                        finalDistance.append(calc_distance_by_icao(jobsFrom[jobIndex],jobsTo[jobIndex]))
                        finalBonusDistance.append("")
                        finalPayPerWeight.append("")
                        finalCosts.append("")
                        finalPayment.append(jobsPay[jobIndex])
                        finalIncome.append(jobsPay[jobIndex])
 
            pandasData = pd.DataFrame({  
            
                '01_From'  :   finalFrom,
                '02_To'  :   finalTo,
                '03_FromName'  :   finalFromName,
                '04_ToName'  :   finalToName,
			    '05_distance' : finalDistance,
                '06_income' : finalIncome
                            })
            pandasData.to_excel(pandasWriter, sheet_name = acfSheetName)                        


        else:
            for curAcfIndex in range(0, len(listLocation)):
                print(acfSheetName + ": " + listLocation[curAcfIndex])

                bestFrom = []
                bestTo = []
                bestAmount = []
                bestUnit = []
                bestPay = []
                bestGreen = []
                bestType = []
                bestAircraft = []

                bestWeight = []
                bestPayPerWeight = []
                bestDistance = []
                bestBonusDistance = []
                
                bestIncome = -1
                bestCosts = 0
                bestPayment = 0
                
                destFound = True
                while destFound == True:
                    destFound = False
                    destination = ""
                    for jobIndex in range(0, len(jobsFrom)):
                        if (jobsFrom[jobIndex] == listLocation[curAcfIndex]) and (jobsType[jobIndex] == "Trip-Only") :
                            if destination == "": destination = jobsTo[jobIndex]
                            destFound = True;
                   
                    if destFound == True:
                        #if (bestIncome == -1): print("Something Found")

                        #print(destination)
                        cargoLeft = listCargoSpace[curAcfIndex]
                        selectedPayment = 0
                        selectedCosts = 0
                        selectedIncome = 0
                        selectedRange = 0
                        crew = acfCrew[acfIndex]
                        seatsLeft = acfSeats[acfIndex] - crew - 1
                        speed = acfSpeed[acfIndex]
                        gph = acfGPH[acfIndex]
                        greenCount = 0
                        greenIncome = 0
                        rental = listRentalDry[curAcfIndex]
                        bonus = listBonus[curAcfIndex]
                        homeBase = listHome[curAcfIndex]
                        # Calc Range Here
                        selectedRange = calc_distance_by_icao(listLocation[curAcfIndex],destination)
                        flightTime = selectedRange / (speed * speedFactor)

                        # Rental Costs
                        selectedCosts += flightTime * rental

                        #print("Costs after Rental: " + str(selectedCosts))
                        # Fuel Costs
                        selectedCosts += flightTime * gph * 4
                        #print("Costs after Fuel: " + str(selectedCosts))
                        
                        # Calc Bonus Range here
                        bonusRange = calc_distance_by_icao(listHome[curAcfIndex], destination) - calc_distance_by_icao(listHome[curAcfIndex], listLocation[curAcfIndex])
                        # Bonus Costs
                        selectedCosts += bonusRange * bonus / 100
                        #print("Costs after Bonus: " + str(selectedCosts))

                        # Additional Crew Fee
                        selectedCosts += flightTime * 100 * crew
                        #print("Costs after Crew: " + str(selectedCosts))

                        selectedFrom = []
                        selectedTo = []
                        selectedAmount = []
                        selectedUnit = []
                        selectedPay = []
                        selectedGreen = []
                        selectedType = []
                        selectedAircraft = []
                        
                        selectedDistance = []
                        selectedBonusDistance = []

                        selectedWeight = []
                        selectedPayPerWeight = []



                        searchAgain = True
                        while searchAgain == True:
                            jobMaxIndex = -1
                            ppw_max = -1

                            for jobIndex in range(0, len(jobsFrom)):
                                if (jobsFrom[jobIndex] == listLocation[curAcfIndex]) and (jobsTo[jobIndex] == destination) and (jobsPayPerWeight[jobIndex] > ppw_max) and (jobsType[jobIndex] == "Trip-Only") :
                                    
                                    ppw_max = jobsPayPerWeight[jobIndex]
                                    jobMaxIndex = jobIndex;

                            if (jobMaxIndex > -1):
                                #print("Here we go")
                                if (jobsWeight[jobMaxIndex] <= cargoLeft) and ((jobsUnit[jobMaxIndex] == "kg") or ((jobsUnit[jobMaxIndex] == "passengers") and (seatsLeft > jobsAmount[jobMaxIndex]))):
                                    #print("YES")
                                    cargoLeft -= jobsWeight[jobMaxIndex]
                                    if (jobsUnit[jobMaxIndex] == "passengers"): seatsLeft -= jobsAmount[jobMaxIndex]
                                    if (jobsGreen[jobMaxIndex] == True): 
                                        greenCount += 1
                                        greenIncome += jobsPay[jobMaxIndex]
                                    selectedPayment += jobsPay[jobMaxIndex]

                                    selectedFrom.append( jobsFrom.pop(jobMaxIndex))
                                    selectedTo.append( jobsTo.pop(jobMaxIndex))
                                    selectedAmount.append( jobsAmount.pop(jobMaxIndex))
                                    selectedUnit.append( jobsUnit.pop(jobMaxIndex))
                                    selectedPay.append( jobsPay.pop(jobMaxIndex))
                                    selectedGreen.append( jobsGreen.pop(jobMaxIndex))
                                    selectedType.append( jobsType.pop(jobMaxIndex))
                                    selectedAircraft.append( jobsAircraft.pop(jobMaxIndex))
                                    selectedDistance.append(selectedRange)
                                    selectedBonusDistance.append(bonusRange)
                                    selectedWeight.append( jobsWeight.pop(jobMaxIndex))
                                    selectedPayPerWeight.append( jobsPayPerWeight.pop(jobMaxIndex))
                                    #print(str(len(selectedFrom)))
                                    
                                    #print("Found, remaining:" + str(len(jobsFrom)))

                                else:
                                    #print("NO")
                                    jobsFrom.pop(jobMaxIndex)
                                    jobsTo.pop(jobMaxIndex)
                                    jobsAmount.pop(jobMaxIndex)
                                    jobsUnit.pop(jobMaxIndex)
                                    jobsPay.pop(jobMaxIndex)
                                    jobsGreen.pop(jobMaxIndex)
                                    jobsType.pop(jobMaxIndex)
                                    jobsAircraft.pop(jobMaxIndex)

                                    jobsWeight.pop(jobMaxIndex)
                                    jobsPayPerWeight.pop(jobMaxIndex) 
                                    #print("Not usable, remaining:" + str(len(jobsFrom)))
                            else:
                                # Route Finished   
                                searchAgain = False
                                # Green Fee
                                if (greenCount > 5):
                                    selectedCosts += greenIncome * greenCount / 100.0;
                                    #print("Costs after Green Fee: " + str(selectedCosts))

                                # Booking Fee
                                selectedCosts += selectedPayment * 0.1
                                #print("Costs after Ground Crew: " + str(selectedCosts))

                                selectedIncome = selectedPayment - selectedCosts

                                if (selectedIncome > bestIncome):
                                    bestIncome = selectedIncome
                                    bestCosts = selectedCosts
                                    bestPayment = selectedPayment

                                    bestFrom =            selectedFrom        
                                    bestTo =              selectedTo 
                                    bestAmount =          selectedAmount 
                                    bestUnit =            selectedUnit 
                                    bestPay =             selectedPay 
                                    bestGreen =           selectedGreen 
                                    bestType =            selectedType 
                                    bestAircraft =        selectedAircraft 
                                    
                                    bestDistance =        selectedDistance
                                    bestBonusDistance =   selectedBonusDistance

                                    bestWeight =          selectedWeight 
                                    bestPayPerWeight =    selectedPayPerWeight 
                                #else:
                                    #print("Bad Income:" + str(selectedIncome))

                if bestIncome > 0:
                    for incomeIndex in range(0,len(bestFrom)):
                        finalFrom.append(                bestFrom[incomeIndex])
                        finalTo.append(                  bestTo[incomeIndex])
                        finalFromName.append(            get_icao_name(bestFrom[incomeIndex]))
                        finalToName.append(              get_icao_name(bestTo[incomeIndex]))
                        finalAmount.append(              bestAmount[incomeIndex])
                        finalUnit.append(                bestUnit[incomeIndex])
                        finalPay.append(                 bestPay[incomeIndex])
                        finalGreen.append(               bestGreen[incomeIndex])
                        finalType.append(                bestType[incomeIndex])
                        finalAircraft.append(            bestAircraft[incomeIndex])

                        finalWeight.append(              bestWeight[incomeIndex])
                        finalPayPerWeight.append(        bestPayPerWeight[incomeIndex])
                        
                        finalDistance.append(            bestDistance[incomeIndex])
                        finalBonusDistance.append(       bestBonusDistance[incomeIndex])


                        finalCosts.append(bestCosts);
                        finalPayment.append(bestPayment);
                        finalIncome.append(bestIncome);
                else:
                    print("Nothing Found")



            pandasData = pd.DataFrame({  
            
                '01_From'  :   finalFrom,
                '02_To'  :   finalTo,
                '03_FromName'  :   finalFromName,
                '04_ToName'  :   finalToName,
                '05_Amount'  :   finalAmount,
                '06_Unit'  :   finalUnit,
                '07_Pay'  :   finalPay,
                '08_Green'  :   finalGreen,
                '09_Type'  :   finalType,
                '10_Aircraft'  :   finalAircraft,

                '11_Weight'  :   finalWeight,
                '12_PayPerWeight'  :   finalPayPerWeight  ,
			    '13_bonusDistance' : finalBonusDistance,
			    '14_distance' : finalDistance,
                '15_costs' : finalCosts,
                '16_payment' : finalPayment,
                '17_income' : finalIncome
                            })
            pandasData.to_excel(pandasWriter, sheet_name = acfSheetName)
        

    else:
        print("Error: Index not found")
		
pandasWriter.save()